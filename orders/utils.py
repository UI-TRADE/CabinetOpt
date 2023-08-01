import os
from django.conf import settings
from django.contrib.staticfiles import finders

from num2words import num2words
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy


def get_order_settings(context):
    return {
        'Заголовок'           : f'Заказ клиента № {context["order"].pk} от {context["order"].created_at.strftime("%d.%m.%Y")}',
        'Поставщик'           : 'ИНН 7802669141, КПП 780201001, ЮИ-ТРЕЙД ООО, 194100, Санкт-Петербург г, Кантемировская ул, дом № 5, корпус 8, комната 104',
        'Покупатель'          : f'ИНН { context["order"].client.inn }, { context["order"].client.name }, { context["contact_detail"].get_address() }, тел.: { context["order"].client.registration_order.phone }',
        'ПодвалСуммаБезСкидки': context["order_totals"]["total_sum_without_discount"],
        'ПодвалСкидка'        : context["order_totals"]["total_discount"],
        'ПодвалСумма'         : context["order_totals"]["total_sum"],
        'ПодвалНДС'           : 'Без НДС',
        'ПодвалКоличество'    : f'Всего наименований { context["order_totals"]["total_count"] }, на сумму { context["order_totals"]["total_sum"] } руб.',
        'ПодвалПропись'       : f'{ in_words(context["order_totals"]["total_sum"]) } руб. {str(context["order_totals"]["total_sum"])[-2:]} коп.'.capitalize()
    }  


def get_line_settings(order_items):
    start_line = 11
    item_titles = {
        'num'                 : '№',
        'product'             : 'Товар',
        'quantity'            : 'Кол-во',
        'weigth'              : 'Вес',
        'unit'                : 'Ед.',
        'price'               : 'Цена',
        'sum_without_discount': 'Сумма без скидки',
        'discount'            : 'Скидка',
        'sum'                 : 'Сумма'
    }

    result = []
    for num, order_item in enumerate(order_items, start=1):
        result.append({
            'num'                 : num,
            'product'             : order_item.product,
            'quantity'            : order_item.quantity,
            'weigth'              : order_item.weight,
            'unit'                : order_item.unit,
            'price'               : order_item.price,
            'sum_without_discount': order_item.get_cost_without_discount(),
            'discount'            : order_item.discount,
            'sum'                 : order_item.sum
        })

    return result, item_titles, start_line    


def in_words(numder):
    if numder:
        return num2words(numder, lang='ru')
    return ''


def get_all_merged_cells(sheet):
    result = []
    for merged_range in sheet.merged_cells:
        merged_cells = {'coords': []}
        coords_of_merged, *_ = sheet[merged_range.coord]
        for coord_of_merged in coords_of_merged:
            merged_cells['coords'].append({'column': coord_of_merged.column, 'row': coord_of_merged.row})
        result.append(merged_cells)
    return result


def find_merged_coords(merged_coords, column, row):
    for merged_cells in merged_coords:
        found_coords = [(item['column'], item['row']) for item in merged_cells['coords'] if item['column'] == column and item['row'] == row]
        if found_coords:
            first_cell, *_, end_cell = merged_cells['coords']
            return first_cell, end_cell
        

def filter_merged_cells(all_merged_cells, selected_cells):
    merged_cells = []
    for row in selected_cells:
        for cell in row:
            if not isinstance(cell, MergedCell):
                continue
            first_cell, end_cell = find_merged_coords(all_merged_cells, cell.column, cell.row)
            merged_cells.append((
                (first_cell['column'], first_cell['row']),
                (end_cell['column'], first_cell['row'])
            ))
    return list(set(merged_cells))


def move_merged_cells(sheet, merged_cells, row_step=1):

    for idx, item in enumerate(merged_cells):

        start_cell, end_cell    = item
        start_column, start_row = start_cell
        end_column, end_row     = end_cell
        new_start_row           = start_row+row_step
        new_end_row             = end_row+row_step
        sheet.unmerge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
        sheet.merge_cells(
            start_row=new_start_row, start_column=start_column,
            end_row=new_end_row, end_column=end_column
        )
        merged_cells[idx] = ((start_column, new_start_row), (end_column, new_end_row))


def save_xlsx(context, response):
    filepath = os.path.join(finders.find(''), 'template.xlsx')
    if not filepath:
        return

    defined_names = get_order_settings(context)
    order_items, item_titles, start_row = get_line_settings(context["order_items"])

    wb = load_workbook(filepath)
    sheet = wb.active

    for key, value in defined_names.items():
        found_range = wb.defined_names[key]
        dests = found_range.destinations
        for _, coord in dests:
            cell = sheet[coord]
            cell.value = value

    last_row = sheet.max_row+1
    last_column = get_column_letter(sheet.max_column-1)
    all_merged_cells = get_all_merged_cells(sheet)
    merged_cells = filter_merged_cells(
        all_merged_cells,
        sheet[f'B{start_row}:{last_column}{last_row}']
    )

    line_area = wb.defined_names['Строка']
    for _, coord in line_area.destinations:
        cell_range, *_ = sheet[coord]
        selected_cells = [cell for cell in cell_range]

        for idx, order_item in enumerate(order_items, start=start_row):

            move_merged_cells(sheet, merged_cells)
            sheet.move_range(f'B{idx}:{last_column}{last_row}', rows=1)

            for source_cell in selected_cells:

                title_cell  = sheet.cell(row = 9, column = source_cell.column)
                target_cell = sheet.cell(row = idx, column = source_cell.column)
                if title_cell.value:
                    for key, value in item_titles.items():
                        if value == title_cell.value:
                            target_cell.value         = str(order_item[key])

                target_cell.font          = copy(source_cell.font)
                target_cell.alignment     = copy(source_cell.alignment)
                target_cell.border        = copy(source_cell.border)
                target_cell.fill          = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection    = copy(source_cell.protection)

                if isinstance(source_cell, MergedCell):
                    first_cell, end_cell = find_merged_coords(all_merged_cells, source_cell.column, source_cell.row)
                    if end_cell['column'] == source_cell.column:
                        sheet.merge_cells(
                            start_row=idx, start_column=first_cell['column'],
                            end_row=idx, end_column=source_cell.column
                        )
            sheet.row_dimensions[idx].height = sheet.row_dimensions[start_row-1].height
            last_row += 1

        sheet.row_dimensions[start_row-1].hidden = True

    wb.save(response)


def read_xlsx(file_path):
    result = []
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        result.append(row)

    return result[1:]
